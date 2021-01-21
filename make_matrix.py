import numpy as np
import openpyxl
from openpyxl.utils import get_column_letter
import os

""" CONSTS """
PATH = os.path.dirname(os.path.abspath(__file__)) + '\\'


def find_excel_file():
    return openpyxl.load_workbook(PATH + '\\input\\' + os.listdir(path=f'{PATH}\\input')[0])


def clear_folder():
    for elem in enumerate(os.listdir(path=f'{PATH}\\output')):
        print(elem)
        os.remove(PATH + '\\output\\' + elem[1])


def make_matrix_a():
    input_wb = find_excel_file()
    sheet = input_wb[input_wb.sheetnames[0]]
    make_matrix(input_wb, sheet)


def make_matrix_b(iterations=None, amount=None):
    input_wb = find_excel_file()
    sheet = input_wb[input_wb.sheetnames[0]]
    for i in range(0, iterations):
        if i == 0:
            make_matrix(input_wb, sheet, b=True, Y=i)
        else:
            if X < sheet.max_row:
                sheet.delete_rows(3, amount=amount)
                make_matrix(input_wb, sheet, b=True, Y=i)
            else:
                break


def make_matrix(input_wb, sheet, b=None, Y=None):
    matrix = np.zeros((sheet.max_row - 2, sheet.max_column - 1))
    for i in range(2, sheet.max_column + 1):
        for j in range(3, sheet.max_row + 1):
            matrix[j - 3][i - 2] = sheet.cell(row=j, column=i).value
    matrix = np.dot(matrix.transpose(), matrix)
    output_wb = openpyxl.Workbook()
    output_wb.title = "Bert Matrix"
    output_sheet = output_wb[output_wb.sheetnames[0]]
    for i in range(1, sheet.max_column + 1):
        output_sheet.cell(row=i, column=1).value = sheet.cell(row=2, column=i).value
        column_widths = len(sheet.cell(row=2, column=i).value)
        output_sheet.column_dimensions[get_column_letter(i)].width = column_widths * 1.5
    for i in range(1, sheet.max_column + 1):
        output_sheet.cell(row=1, column=i).value = sheet.cell(row=2, column=i).value
        column_widths = len(sheet.cell(row=2, column=i).value)
        output_sheet.column_dimensions[get_column_letter(i)].width = column_widths * 1.5
    for i in range(2, sheet.max_column + 1):
        for j in range(2, sheet.max_column + 1):
            output_sheet.cell(row=i, column=j).value = matrix[i - 2][j - 2]
    if b:
        output_wb.save(filename=PATH + f'\\output\\B_{Y}' + os.listdir(path=f'{PATH}\\input')[0])
    else:
        output_wb.save(filename=PATH + '\\output\\A_' + os.listdir(path=f'{PATH}\\input')[0])


if __name__ == '__main__':
    clear_folder()
    print('A - Make Bert Matrix. \nB - Make Bert Matrix but delete X values in each Y iteration')
    type_work = input('>> ')
    if type_work.strip() == 'A':
        make_matrix_a()
    if type_work.strip() == 'B':
        print('Enter X and Y through a space.')
        X, Y = input('').split(' ')
        try:
            X, Y = int(X), int(Y)
            make_matrix_b(Y, X)
        except ValueError:
            print('ValueError. Invalid X or Y')


